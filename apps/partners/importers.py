"""Импорт контрагентов из Excel (.xlsx).

Поддерживает два формата:
1. Экспорт контрагентов из «МойСклад» (полный набор колонок, включая
   банковские реквизиты — из них создаётся банковский счёт).
2. Простой шаблон: Наименование | Форма | Тип | ИНН | КПП | Телефон | E-mail | Адрес | Контактное лицо

Строка заголовков ищется в первых 10 строках каждого листа —
файл с пустыми строками сверху тоже импортируется.
Существующие контрагенты обновляются по совпадению ИНН (без ИНН — по наименованию).
"""
from django.db import transaction
from openpyxl import load_workbook

from .models import BankAccount, Counterparty

KIND_MAP = {
    "юл": Counterparty.KIND_LEGAL,
    "юридическое лицо": Counterparty.KIND_LEGAL,
    "ип": Counterparty.KIND_ENTREPRENEUR,
    "индивидуальный предприниматель": Counterparty.KIND_ENTREPRENEUR,
    "физлицо": Counterparty.KIND_PERSON,
    "физическое лицо": Counterparty.KIND_PERSON,
}
TYPE_MAP = {
    "покупатель": Counterparty.TYPE_CUSTOMER,
    "поставщик": Counterparty.TYPE_SUPPLIER,
    "оба": Counterparty.TYPE_BOTH,
    "покупатель и поставщик": Counterparty.TYPE_BOTH,
}


def _clean(value):
    if value is None:
        return ""
    return str(value).strip()


def _digits_only(value):
    """Для счетов и БИК: убирает пробелы и прочие разделители — остаются цифры."""
    return "".join(ch for ch in _clean(value) if ch.isdigit())


def _truncate_fields(obj):
    """Обрезает строковые поля до max_length модели (PostgreSQL иначе падает)."""
    for f in obj._meta.concrete_fields:
        max_length = getattr(f, "max_length", None)
        if max_length:
            value = getattr(obj, f.attname, None)
            if isinstance(value, str) and len(value) > max_length:
                setattr(obj, f.attname, value[:max_length])


def _find_header(ws, max_rows=10):
    """Ищет строку заголовков (содержит «Наименование»). Возвращает (номер строки, {колонка: индекс})."""
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows, values_only=True), start=1):
        cols = {_clean(h).lower(): i for i, h in enumerate(row) if _clean(h)}
        if "наименование" in cols:
            return row_idx, cols
    return None, None


def _parse_kind(raw):
    """«Юридическое лицо. Россия» (МойСклад) или «ЮЛ» (простой шаблон)."""
    low = raw.lower()
    for key, kind in KIND_MAP.items():
        if low.startswith(key):
            return kind
    return None


def import_counterparties(file):
    # НЕ read_only: экспорт МойСклада объявляет неверную размерность листа,
    # и в экономном режиме openpyxl обрезает строки до одной ячейки (теряются
    # все колонки, кроме первой). Обычный режим читает реальные ячейки.
    wb = load_workbook(file, data_only=True)

    header_row, col = None, None
    ws = None
    for sheet in wb.worksheets:
        header_row, col = _find_header(sheet)
        if col:
            ws = sheet
            break
    if not col:
        found = ", ".join(_clean(c) for c in next(wb.active.iter_rows(max_row=1, values_only=True), []) if _clean(c))
        return 0, 0, [
            "Не найдена колонка «Наименование» ни на одном листе. "
            f"Первая строка активного листа: {found or 'пустая'}",
        ]

    def cell(row, *names):
        for name in names:
            idx = col.get(name)
            if idx is not None and idx < len(row):
                value = _clean(row[idx])
                if value:
                    return value
        return ""

    created = updated = 0
    errors = []

    for n, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        name = cell(row, "наименование")
        if not name:
            continue
        try:
            # Каждая строка — в своей транзакции: ошибка одной не ломает остальные
            # (на PostgreSQL сбойная операция иначе рушит всю транзакцию запроса).
            with transaction.atomic():
                was_created = _import_row(cell, row, name)
            created += was_created
            updated += not was_created
        except Exception as exc:  # noqa: BLE001 — одна плохая строка не роняет импорт
            errors.append(f"Строка {n}: {exc}")

    return created, updated, errors


def _import_row(cell, row, name):
    """Создаёт/обновляет одного контрагента из строки. Возвращает was_created."""
    inn = cell(row, "инн")
    obj = None
    if inn:
        obj = Counterparty.objects.filter(inn=inn).first()
    if obj is None:
        obj = Counterparty.objects.filter(name=name, inn="").first()
    was_created = obj is None
    if was_created:
        obj = Counterparty(name=name)

    obj.name = name
    obj.inn = inn or obj.inn
    # заполняем только присутствующие в файле колонки — не затираем остальное
    field_map = [
        ("full_name", ("полное наименование",)),
        ("kpp", ("кпп",)),
        ("okpo", ("окпо",)),
        ("phone", ("телефон",)),
        ("email", ("e-mail", "email", "электронный адрес")),
        ("legal_address", ("юридический адрес", "адрес")),
        ("actual_address", ("фактический адрес",)),
        ("contact_person", ("контактное лицо",)),
        ("comment", ("комментарий",)),
    ]
    for field, names in field_map:
        value = cell(row, *names)
        if value:
            setattr(obj, field, value)

    ogrn = cell(row, "огрн") or cell(row, "огрнип")
    if ogrn:
        obj.ogrn = ogrn

    kind = _parse_kind(cell(row, "тип контрагента", "форма"))
    if kind:
        obj.kind = kind

    ptype = TYPE_MAP.get(cell(row, "тип").lower())
    if ptype:
        obj.partner_type = ptype

    archived = cell(row, "архивный").lower()
    if archived in ("да", "yes", "true", "1"):
        obj.is_active = False
    elif archived in ("нет", "no", "false", "0"):
        obj.is_active = True

    _truncate_fields(obj)
    obj.save()

    # Банковский счёт: номера и БИК чистим от пробелов, длинное — обрезаем
    account = _digits_only(cell(row, "р/с", "расчётный счёт", "расчетный счет"))
    if account:
        BankAccount.objects.update_or_create(
            counterparty=obj, account=account[:20],
            defaults={
                "bank_name": cell(row, "банк")[:255],
                "bik": _digits_only(cell(row, "бик"))[:9],
                "corr_account": _digits_only(cell(row, "к/с", "корр. счёт", "корр. счет"))[:20],
            },
        )

    return was_created
