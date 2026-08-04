"""Импорт товаров из Excel (.xlsx).

Ожидаемые колонки в первой строке (порядок не важен, регистр не важен):
Наименование*, Тип (Товар/Услуга), Группа, Артикул, Код, Штрихкод,
Ед. изм., Ставка НДС (20/10/0/Без НДС), Закупочная цена, Цена продажи, Мин. остаток
"""
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook

from apps.core.constants import VAT_0, VAT_10, VAT_20, VAT_NONE

from .models import Product, ProductGroup, Unit

VAT_MAP = {
    "20": VAT_20, "20%": VAT_20,
    "10": VAT_10, "10%": VAT_10,
    "0": VAT_0, "0%": VAT_0,
    "без ндс": VAT_NONE, "нет": VAT_NONE, "-": VAT_NONE,
}


def _dec(value, default=Decimal("0")):
    if value in (None, ""):
        return default
    try:
        return Decimal(str(value).replace(",", ".").replace(" ", ""))
    except InvalidOperation:
        return default


def import_products(file):
    wb = load_workbook(file, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)

    header = next(rows, None)
    if not header:
        return 0, 0, ["Файл пуст"]
    col = {str(h).strip().lower(): i for i, h in enumerate(header) if h}

    def cell(row, name):
        idx = col.get(name)
        if idx is None or idx >= len(row):
            return ""
        return str(row[idx]).strip() if row[idx] is not None else ""

    if "наименование" not in col:
        return 0, 0, ["Не найдена колонка «Наименование»"]

    default_unit, _ = Unit.objects.get_or_create(name="шт", defaults={"okei_code": "796"})
    created = updated = 0
    errors = []

    for n, row in enumerate(rows, start=2):
        name = cell(row, "наименование")
        if not name:
            continue
        try:
            unit_name = cell(row, "ед. изм.") or cell(row, "ед.изм.") or cell(row, "единица")
            unit = Unit.objects.get_or_create(name=unit_name)[0] if unit_name else default_unit

            group = None
            group_name = cell(row, "группа")
            if group_name:
                group, _ = ProductGroup.objects.get_or_create(name=group_name, parent=None)

            article = cell(row, "артикул")
            item_type = Product.TYPE_SERVICE if cell(row, "тип").lower() == "услуга" else Product.TYPE_PRODUCT
            vat = VAT_MAP.get(cell(row, "ставка ндс").lower().rstrip(".0") or "20", VAT_20)

            defaults = {
                "item_type": item_type,
                "group": group,
                "code": cell(row, "код"),
                "barcode": cell(row, "штрихкод"),
                "unit": unit,
                "vat_rate": vat,
                "purchase_price": _dec(row[col["закупочная цена"]] if "закупочная цена" in col else None),
                "sale_price": _dec(row[col["цена продажи"]] if "цена продажи" in col else None),
                "min_stock": _dec(row[col["мин. остаток"]] if "мин. остаток" in col else None),
            }
            if article:
                obj, was_created = Product.objects.update_or_create(
                    article=article, defaults={"name": name, **defaults},
                )
            else:
                obj, was_created = Product.objects.update_or_create(
                    name=name, article="", defaults=defaults,
                )
            created += was_created
            updated += not was_created
        except Exception as exc:  # noqa: BLE001 — одна плохая строка не должна ронять импорт
            errors.append(f"Строка {n}: {exc}")

    return created, updated, errors
